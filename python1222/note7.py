

'''
1、excel测试用例准备ok,代码可以自动读取用例数据------read_case(filename,sheetname)
2、执行接口测试，得到响应结果-------api_fun(url,data)
3、断言：响应结果==预期结果   ----通过/不通过？
4、写到 最终执行通过与否的结果-----write_result(filename,sheetname,row,column,final_result)

'''


import openpyxl
import requests
import jsonpath

# 读取测试用例
def read_case(filename,sheetname):
    wb = openpyxl.load_workbook(filename)   # 加载工作簿，打开一个excel文件
    sheet = wb[sheetname]                              # 打开某一个表单
    row_max = sheet.max_row                             # 获取最大行数
    case_list = []                             # 新建空列表，存放for循环依次读取到的测试用例数
    for i in range(2,row_max+1):
        data_dict = dict(
            case_id = sheet.cell(row=i,column=1).value,
            url = sheet.cell(row=i,column=5).value,          # 读取url值
            data = sheet.cell(row=i,column=6).value,         # 读取data值
            expected = sheet.cell(row=i,column=7).value      # 读取期望值
            )
        case_list.append(data_dict)        # 把每一个读取到的测试用例数据生成否如字典，追加到list中
    return case_list


# 执行测试用例
def api_fun(url,data):
    headers = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}

    res = requests.post(url=url,json=data,headers=headers).json()
    return res

# 写入测试结果
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)     # 加载工作簿，打开一个excel文件
    sheet = wb[sheetname]                     # 打开某一个表单
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)                         # 保存



# 将这整个过程封装，定义成一个函数：
def execute_fun(filename,sheetname):
    # 调用函数读取测试用例：
    cases = read_case(filename,sheetname)    # 调用函数，设置变量接收返回值
    for case in cases:                                    # ---此处获取的都是str格式
        case_id = case['case_id']
        url = case['url']                         # 获取url的值
        data = eval(case['data'])                 # 获取data的值,使用eval运行出里面的dict格式
        expected = eval(case['expected'])         # 获取期望值,使用eval运行出里面的dict格式
        expect_msg = expected['msg']              # 获取预期中的msg

    # eval（） --- 运行被字符串包裹着的表达式
    # '{"mobile_phone": "18362980328", "pwd": "lemon666", "type":"1","reg_name":"小安"}'  # 字符串
    # dict0 = eval('{"mobile_phone": "18362980328", "pwd": "lemon666", "type":"1","reg_name":"小安"}')
    # print(dict0)
    # print(type(dict0))
    # print(eval('2+6'))

    # 执行测试用例：
        real_result = api_fun(url,data)                    # 调用函数，执行测试用例
        real_msg = real_result['msg']
        print(case_id,expect_msg,real_msg)

    # 断言
        if expect_msg == real_msg:
            print('第{}条用例执行通过'.format(case_id))
            final_re = 'passed'
        else:
            print('第{}条用例执行不通过'.format(case_id))
            final_re = 'failed'
        print('*'*20)                                       # 分隔增加阅读性

    # 写入测试结果
        write_result(filename,sheetname,case_id+1,8,final_re)

# 调用这个函数：
execute_fun('../test_data/test_case_api.xlsx', 'register')
execute_fun('../test_data/test_case_api.xlsx', 'login')





